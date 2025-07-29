from sqlite3.dbapi2 import paramstyle

import pandas as pd
import json
import numpy as np
from openpyxl import load_workbook
import os


excel_file = r'C:\Users\JNewton\Documents\GitHub\portfolio-sensitivity-prototype\examples\Sensitivity Set Up.xlsx'
output_json = r'C:\Users\JNewton\Documents\GitHub\portfolio-sensitivity-prototype\examples\sensitivity_set_up.json'
technologies = ["solar"]
folder = "3315"

def get_sweep_type(file_path: str) -> str:
    wb = load_workbook(file_path, read_only=True, data_only=True)

    sh = wb[list(wb.sheetnames)[0]]  # Get the first sheet
    if not sh:
        raise ValueError("No sheets found in the Excel file.")
    sweep_type = sh.cell(row=3, column=2).value
    if sweep_type not in ["Independent", "Linked"]:
        raise ValueError(f"Invalid sweep type: {sweep_type}. Expected 'Independent' or 'Linked'.")
    print(f"Sweep type: {sweep_type}")
    return sweep_type

def read_excel_file(file_path: str) -> pd.DataFrame:

    df = pd.read_excel(file_path, header=4, index_col=0, nrows=8, engine='openpyxl', usecols="B:G" )
    #rename Min Value and Max Value to Min and Max
    df = df.rename(columns={'Min Value': 'Min', 'Max Value': 'Max'})

    return df

def df_to_json(df: pd.DataFrame) -> None:
    """Convert DataFrame to JSON and save it to the specified path."""

    # drop rows where Include is False
    sweep_df = df[df['Include']]
    # drop the Include column
    sweep_df = sweep_df.drop(columns=['Include'])

    # for each roe, create a dictionary with parameter and list of values
    sweep_data = {}
    for index, row in sweep_df.iterrows():
        parameter = index
        type = row['Adjustment Type'].lower()
        min_value = row['Min']
        max_value = row['Max']
        step = row['Step']
        value_range = np.arange(min_value, max_value +step, step)
        value_range = np.round(value_range, 6)

        sweep_data[parameter] = {
            "component": parameter,
            "type": type,
            "values": list(value_range)
        }

        if get_sweep_type(excel_file) == "Independent":

            sensitivity = {
                "folder": folder,
                "technologies": technologies,

                "sensitivities": {
                }
            }

            for sweeps in sweep_data.values():
                sensitivity["sensitivities"]["%s_sweep" % sweeps["component"]] = {"element_wise_parameter_sweep":{
                    parameter:{
                    "component": sweeps["component"],
                    "type": sweeps["type"],
                    "values": sweeps["values"]
                        }
                    }
                }

        elif get_sweep_type(excel_file) == "Linked":
            sensitivity = {
                "folder": folder,
                "technologies": technologies,

                "sensitivities": {
                    "element_wise_sweep": {
                        "element_wise_parameter_sweep": {}
                    }
                }
            }

            for sweeps in sweep_data.values():
                sensitivity["sensitivities"]["element_wise_sweep"]["element_wise_parameter_sweep"]["%s"  % sweeps["component"]] = {
                    "component": sweeps["component"],
                    "type": sweeps["type"],
                    "values": sweeps["values"]
                }

    # Save the JSON data to a file
    with open(output_json, 'w') as f:
        json.dump(sensitivity, f, indent=4)
    print(f"JSON file created at {output_json}.")

if __name__ == "__main__":
    try:
        df = read_excel_file(excel_file)
    except:
        print(f"Error reading the Excel file: {excel_file}")
        raise
    sweep_data = df_to_json(df)