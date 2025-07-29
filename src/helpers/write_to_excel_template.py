import logging
import os
import time
from collections.abc import Callable
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from src.helpers.format_time_taken import format_time_taken
from src.models.enums.sensitivities import ScenarioComponents
from src.models.gem_assessments import IndividualSensitivityResult, SensitivityResults

TEMPLATE_EXCEL = os.path.join("templates", "sensitivity_template_v1.xlsx")

TABLE_NAME = "SensitivityResults"

FieldAccessor = Callable[[IndividualSensitivityResult], Any]


PORTFOLIO_MAPPER = {
    "England": "UK&I",
    "Scotland": "UK&I",
    "Wales": "UK&I",
    "Northern Ireland": "UK&I",
    "Republic of Ireland": "UK&I",
    "Sweden": "Sweden",
    "Norway": "Norway",
    "Germany": "Germany",
    "Türkiye": "Türkiye",
    "New South Wales": "Australia",
    "Victoria": "Australia",
    "Queensland": "Australia",
    "Western Australia": "Australia",
    "South Australia": "Australia",
    "Tasmania": "Australia",
    "Northern Territory": "Australia",
    "Texas": "USA",
}

TECHNOLOGY_MAPPER = {
    "WIND": "Onshore Wind",
    "SOLAR": "Solar",
    "STORAGE": "Battery Storage",
}

FIELD_MAPPING: dict[str, FieldAccessor] = {
    "Project ID": lambda r: r.project_id,
    "Project Name": lambda r: r.project_name,
    "Phase": lambda r: r.phase,
    "Technology": lambda r: TECHNOLOGY_MAPPER[r.technology],
    "Portfolio": lambda r: PORTFOLIO_MAPPER[r.country],
    "Currency": lambda r: r.currency,
    "Scenario Name": lambda r: r.scenario,
    "Discount Rate Adjustment": lambda r: r.combination.get(ScenarioComponents.DISCOUNT_RATE),
    "Capex Adjustment": lambda r: r.combination.get(ScenarioComponents.ALL_CAPEX),
    "Power Prices Adjustment": lambda r: r.combination.get(ScenarioComponents.POWER_PRICES),
    "Opex Adjustment": lambda r: r.combination.get(ScenarioComponents.ALL_OPEX),
    "Lifetime Adjustment": lambda r: r.combination.get(ScenarioComponents.OPERATIONAL_LIFETIME),
    "Financial Close Date Adjustment": lambda r: r.combination.get(ScenarioComponents.FINANCIAL_CLOSE_DATE),
    "Energy Yield Adjustment": lambda r: r.combination.get(ScenarioComponents.ENERGY_YIELD),
    "Total Capex": lambda r: r.results.total_capex if r.results else None,
    "Total Merchant Revenue": lambda r: r.results.total_merchant_revenue if r.results else None,
    "Total Opex": lambda r: r.results.total_opex if r.results else None,
    "Discount Rate": lambda r: r.results.discount_rate if r.results else None,
    "Lifetime": lambda r: r.results.lifetime if r.results else None,
    "Sale Date": lambda r: r.results.project_sale_date if r.results else None,
    "FID": lambda r: r.results.fid if r.results else None,
    "COD": lambda r: r.results.cod if r.results else None,
    "Installed Capacity": lambda r: r.results.installed_capacity if r.results else None,
    "Energy Yield": lambda r: r.results.energy_yield if r.results else None,
    "Development Fee": lambda r: r.results.development_fee if r.results else None,
    "IRR": lambda r: r.results.irr if r.results else None,
    "BEP": lambda r: r.results.bep if r.results else None,
    "Projects with Positive Development Fee": lambda r: 1
    if (r.results and r.results.development_fee >= 0)
    else 0
    if r.results
    else None,
}


def write_results_to_template_excel_file(sensitivity_results: SensitivityResults, output_file: str) -> None:
    start_time = time.time()
    wb = load_workbook(TEMPLATE_EXCEL)
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet found in template.xlsx")

    table: Table | None = next((tbl for tbl in ws.tables.values() if tbl.displayName == TABLE_NAME), None)
    if not table:
        raise ValueError("Table 'SensitivityResults' not found in template.xlsx")

    header_row = int(table.ref.split(":")[0][1:])
    headers = {ws.cell(row=header_row, column=col).value: col for col in range(1, ws.max_column + 1)}

    start_row = header_row + 1

    valid_results = [result for result in sensitivity_results.assessments if (result.reason_for_no_assessment is None)]
    for i, result in enumerate(valid_results, start=start_row):
        if result.reason_for_no_assessment is None:
            for header, func in FIELD_MAPPING.items():
                column = headers.get(header)
                if column:
                    value = func(result)
                    ws.cell(row=i, column=column, value=value)

    end_row = start_row + len(valid_results) - 1
    end_column = get_column_letter(max(headers.values()))
    start_column = get_column_letter(min(headers.values()))
    table.ref = f"{start_column}{header_row}:{end_column}{end_row}"

    wb.save(output_file)
    logging.info(f"File saved as '{output_file}' in {format_time_taken(time.time() - start_time)}")
